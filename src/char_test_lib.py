import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import numpy as np
from scipy import stats

REDFILL, GREENFILL = PatternFill(start_color='F01E2C',end_color='F01E2C',fill_type='solid'),PatternFill(start_color='3BB143',end_color='3BB143',fill_type='solid')

'''
Change default to True to evaluate based on defined maximum and minimum rather than comparing to other scintillators 
'''
acceptance_lib = {
   'Event Rate': {
      'default': True,
      'std_results': True,
      'func': lambda std_results, ch: std_results[ch][3],
      'maximum': None,
      'minimum': None,
      'too_big': 'Check for light leaks (high event rate).',
      'too_small': 'Check for something blocking your scintillators (low event rate).',
   },
   'Energy Overflows': {
      'default': True,
      'std_results': True,
      'func': lambda std_results, ch: std_results[ch][4],
      'maximum': None,
      'minimum': None,
      'too_big': 'High energy saturation.',
      'too_small': 'Low energy saturation.',
   },
   'Langauss MPV': {
      'default': True,
      'std_results': False,
      'func': lambda fit_results, ch: fit_results['Langauss'][ch].params[0],  
      'maximum': None,
      'minimum': None,
      'too_big': 'Larger-than-expected light yield (high mpv of pulse height distribution).',
      'too_small': 'Small light yield or noisy signal (low mpv of pulse height distribution).',
   },
   'Langauss \u03C7\u00B2/ndof': { # chi^2/ndof
      'default': True,
      'std_results': False,
      'func': lambda fit_results, ch: fit_results['Langauss'][ch].test_results[f"chi2/ndof Statistic"],  
      'maximum': None,
      'minimum': None,
      'too_big': 'High \u03C7\u00B2/ndof of Langauss fit (height)',
      'too_small': 'Low \u03C7\u00B2/ndof of Langauss fit (height)',
   },
   'Threshold': {
      'default': True,
      'std_results': False,
      'func': lambda fit_results, ch: fit_results['Langauss'][ch].thresh,  
      'maximum': None,
      'minimum': None,
      'too_big': 'High threshold. Check light yield,',
      'too_small': 'Low threshold. Check light yield and noise levels.',
   },
   'Energy \u03C7\u00B2/ndof': { # chi^2/ndof 
      'default': True,
      'std_results': False,
      'func': lambda fit_results, ch: fit_results['EMG'][ch].test_results[f"chi2/ndof Statistic"],  
      'maximum': None,
      'minimum': None,
      'too_big': '\u03C7\u00B2/ndof of EMG fit (energy)',
      'too_small': '\u03C7\u00B2/ndof of EMG fit (energy)',
   },
   'Energy \u03BC': { # mu
      'default': True,
      'std_results': False,
      'func': lambda fit_results, ch: fit_results['EMG'][ch].params[1],  
      'maximum': None,
      'minimum': None,
      'too_big': 'High \u03BC parameter (energy spectrum)',
      'too_small': 'Low \u03BC parameter (energy spectrum)',
   },
   'Energy \u03C3': { # sigma
      'default': True,
      'std_results': False,
      'func': lambda fit_results, ch: fit_results['EMG'][ch].params[2],  
      'maximum': None,
      'minimum': None,
      'too_big': 'High \u03C3 parameter (energy spectrum)',
      'too_small': 'Low \u03C3 parameter (energy spectrum)',
   },
   'Energy K': {
      'default': True,
      'std_results': False,
      'func': lambda fit_results, ch: fit_results['EMG'][ch].params[0],  
      'maximum': None,
      'minimum': None,
      'too_big': 'High $K$ parameter (energy spectrum)',
      'too_small': 'Low $K$ parameter (energy spectrum)',
   },
}

def calc_accept_param_range(cat_name,  info):
   '''
   Calculates the acceptable parameter range
   `cat_name` - The key of an item contained in `acceptance_lib`
   `info` - A list of dictionaries where the list maps to scintilaltors and the dictionary maps to values of the categories of `acceptance_lib` 
   '''
   cat = acceptance_lib[cat_name]
   minimum, maximum = cat['minimum'], cat['maximum'] # Use defined ranges user has requested
   if cat['default']:
      # Set the acceptable range to be [q1-iqr, q3+iqr] 
      col = [scint[cat_name] for scint in info]
      q1,q3 = np.quantile(col,1/4), np.quantile(col,3/4)
      iqr = q3-q1
      minimum, maximum = q1-iqr*1.5, q3+iqr*1.5
   return minimum, maximum

def grubbs_test(data, potential_outlier, alpha = 0.05):
   '''
   Checks if `potential_outlier` is an outlier in `data` (under the assumption that `data` has no outliers)
   https://www.itl.nist.gov/div898/handbook/eda/section3/eda35h1.htm
   https://www.geeksforgeeks.org/python/how-to-find-the-t-critical-value-in-python/
   '''
   G = np.abs(potential_outlier - np.mean(data)) / np.std(data)

   N = len(data) + 1
   significance_level = alpha / (2 * N)
   t_alpha = stats.t.ppf(1 - significance_level, N - 2)

   rhs = (N-1)/np.sqrt(N) * np.sqrt(t_alpha**2 / (N - 2 + t_alpha**2))
   return G > rhs # We reject the hypothesis of no outliers if G > RHS

def pass_or_fail(outfolder, scint_lst, std_results, fit_results, logger=None):
   '''
   Determines whether or not a scintillator passes or fails and carries out the corresponding behaviour
   '''
   # Calculate test results for each scintilator 
   info,n_scints = [],range(len(scint_lst))
   for ch in n_scints:
      scint_info = {}
      for cat_name, cat in acceptance_lib.items():
         results = std_results if cat['std_results'] else fit_results
         scint_info[cat_name] =  cat['func'](results,ch)
      info.append(scint_info)

   # Initialization of `scintillator_results.xlsx`-based variables
   results_path, n_passed, offset = outfolder / 'scintillator_results.xlsx',0,(0,0)
   if results_path.is_file():
      df = pd.read_excel(results_path,sheet_name='Results')
      pass_mask = df['Success'].to_numpy(dtype=bool)
      n_passed,offset = np.count_nonzero(pass_mask),df.shape

   # Use test results to determine pass/fail status
   ## Calculate acceptable ranges for test results
   maxima, minima = {},{}
   for cat_name in acceptance_lib.keys():
      minima[cat_name], maxima[cat_name] = calc_accept_param_range(cat_name, info)

   ## If there is a sufficient number of passed scintillators already and the user requested default behaviour, use Grubb's test, else use calculated range
   successes,messages = [],[]
   for ch in n_scints:
      success = {}
      message = ''
      for cat_name in acceptance_lib.keys():
         if n_passed > 6 and acceptance_lib[cat_name]['default']:
            data = df[cat_name][pass_mask]
            success[cat_name] = not grubbs_test(data,scint_lst[ch][cat_name])
         else:
            success[cat_name] = not ((scint_lst[ch][cat_name] < minima[cat_name]) or (scint_lst[ch][cat_name] > maxima[cat_name]))
         if not success[cat_name]:
            fail_message = acceptance_lib[cat_name]['too_big'] if scint_lst[ch] > np.mean(data) else acceptance_lib[cat_name]['too_small']
            message += fail_message + '\n'
      successes.append(success)
      messages.append(message)

   # Update `scintillator_results.xlsx`
   ## Create new dataframe
   results = {
      'ID': scint_lst,
      'Success': [sum(list(successes[ch].values()))==len(successes[ch]) for ch in n_scints],
      'Message': [messages[ch] for ch in n_scints],
      'Attempt': [1,1,1,1]
   }
   for cat_name in acceptance_lib.keys():
      results[cat_name] = [scint_lst[ch][cat_name] for ch in n_scints]
   new_df = pd.DataFrame(results)

   ## Open the worksheet
   wb = openpyxl.load_workbook(results_path)
   ws = wb['Results']

   ## Update the worksheet to avoid duplicate scintillators
   if results_path.is_file():
      existing_df = pd.read_excel(results_path,sheet_name='Results')
      to_drop = []
      for row in existing_df.itertuples():
         if row.ID in scint_lst:
            to_drop.append(row.index)
            i=0
            while scint_lst[i] != row.ID: i+=1
            new_df['Attempt'][i] += row.Attempt

      for idx in to_drop.reverse(): ws.delete_rows(idx=idx+1, amount=1)
      offset = existing_df.shape

   ## Append rows
   for row in dataframe_to_rows(new_df, index=False, header=False):
      ws.append(row)

   ## Colour cells based on whether or not they pass or fail
   row_index = 1 + offset[0]
   for ch in n_scints:
      col_index = 1 + offset[1]
      for col_title in results.keys():
         if col_title == 'Success': 
            fill_colour = GREENFILL if results[col_title][ch] else REDFILL
         else:
            fill_colour = GREENFILL if successes[ch][col_title] else REDFILL
         ws.cell(row=row_index, column=col_index).fill = fill_colour
            
   ## Save results
   wb.save(results_path)